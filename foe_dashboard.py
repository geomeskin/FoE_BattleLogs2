"""
FoE Fragment Dashboard
======================
Reads a city JSON export and calculates daily fragment production
for all kits, then shows a dashboard with days remaining per kit.

Usage:
    python foe_dashboard.py <city_json_file> [<city_json_file2> ...] [--excel]

Examples:
    python foe_dashboard.py LUT_Buildings_CTHG_20260407.json
    python foe_dashboard.py LUT_Buildings_CTHG_20260407.json LUT_Buildings_PMTHS_20260330.json --excel
"""

import json
import re
import sys
import os
import pandas as pd
from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
SCRIPT_DIR        = Path(__file__).parent
LUT_DIR           = SCRIPT_DIR / "LUT"
OUTPUT_DIR        = SCRIPT_DIR / "output"
FRAGMENTS_LUT     = LUT_DIR / "LUT_Item_FragmentsNeeded_20260330.csv"
KIT_MAPPING_LUT   = LUT_DIR / "LUT_Kit_Mapping.csv"
HISTORY_FILE      = OUTPUT_DIR / "foe_history.json"
GITHUB_DIR        = SCRIPT_DIR / "github"
ABBREV_LUT        = LUT_DIR / "LUT_aor_FOE_Abbreviations.csv"

CHANGE_THRESHOLD  = 0.1   # minimum delta to show as a change


# ── Helpers ───────────────────────────────────────────────────────────────────

def load_city(json_path: str) -> list[dict]:
    """Load city JSON, handling UTF-8 BOM if present."""
    with open(json_path, encoding="utf-8-sig") as f:
        return json.load(f)


def load_luts() -> tuple[dict, dict]:
    """
    Returns:
        frags_needed  : {normalized_kit_name: fragments_needed}
        kit_mapping   : {UPPERCASE_raw_name: correct_display_name}
    """
    frags_df = pd.read_csv(FRAGMENTS_LUT)
    frags_needed = {}
    for _, row in frags_df.iterrows():
        name = str(row["Item"]).strip()
        needed = int(row["FragmentsNeeded"])
        frags_needed[clean_name(name)] = (name, needed)

    map_df = pd.read_csv(KIT_MAPPING_LUT)
    kit_mapping = {}
    for _, row in map_df.iterrows():
        raw     = clean_name(str(row["Items_clean"]))
        correct = str(row["Correct_LUT_Item"]).strip()
        kit_mapping[raw] = correct

    return frags_needed, kit_mapping


def load_city_names() -> dict:
    """Load abbreviation -> full city name mapping."""
    if not ABBREV_LUT.exists():
        return {}
    try:
        df = pd.read_csv(ABBREV_LUT)
        return {
            str(row["Abbreviation"]).strip(): str(row["City Name"]).strip()
            for _, row in df.iterrows()
        }
    except Exception:
        return {}


def clean_name(name: str) -> str:
    """Normalize a kit name for reliable matching."""
    name = name.strip()
    name = re.sub(r"^Fragments?\s+of\s+", "", name, flags=re.IGNORECASE)
    name = re.sub(r"[\u2010\u2011\u2012\u2013\u2014\u2015\ufe58\ufe63\uff0d]", "-", name)
    name = re.sub(r"[\u2018\u2019\u201a\u201b]", "'", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name.upper()


def normalize_kit_name(raw: str, kit_mapping: dict) -> str:
    """Map a raw kit name to its canonical display name."""
    key = clean_name(raw)
    if key in kit_mapping:
        return kit_mapping[key]
    name = raw.strip()
    name = re.sub(r"^Fragments?\s+of\s+", "", name, flags=re.IGNORECASE)
    name = re.sub(r"[\u2010\u2011\u2012\u2013\u2014\u2015\ufe58\ufe63\uff0d]", "-", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def parse_fragment_entries(text: str) -> list[tuple[float, str]]:
    """Parse all Nx 🧩 KitName entries from a text segment."""
    results = []
    pattern = r"([\d.]+)x\s*🧩\s*(.+?)(?=\s*[\d.]+x\s*🧩|$)"
    matches = re.findall(pattern, text)
    for amt_str, kit_name in matches:
        try:
            amt = float(amt_str)
            kit = re.sub(r"[Ø\s]+$", "", kit_name.strip()).strip()
            if not kit:
                continue
            lone_suffix = re.match(r"^([hm])\s+(.+)$", kit)
            if lone_suffix:
                kit = f"{int(amt)}{lone_suffix.group(1)} {lone_suffix.group(2)}"
                amt = 1.0
            results.append((amt, kit))
        except ValueError:
            continue
    return results


def parse_items_field(items_str: str) -> tuple[list[tuple[float, str]], list[tuple[float, str]]]:
    """Parse the Items field from the city JSON."""
    if not items_str or items_str == "0":
        return [], []

    parts = items_str.split("Ø")
    guaranteed_part = parts[0]
    chance_parts    = parts[1:]

    guaranteed = parse_fragment_entries(guaranteed_part)
    chance     = []
    for part in chance_parts:
        chance.extend(parse_fragment_entries(part))

    return guaranteed, chance


def collect_fragment_production(city: list[dict], kit_mapping: dict) -> tuple[dict, dict]:
    """Walk every building in the city and sum up daily fragment production."""
    guaranteed = {}
    chance     = {}

    for bldg in city:
        count = bldg.get("#", 1)
        items = str(bldg.get("Items", ""))

        guar_entries, chance_entries = parse_items_field(items)

        for amt, raw_kit in guar_entries:
            canonical = normalize_kit_name(raw_kit, kit_mapping)
            guaranteed[canonical] = guaranteed.get(canonical, 0.0) + amt * count

        for amt, raw_kit in chance_entries:
            canonical = normalize_kit_name(raw_kit, kit_mapping)
            chance[canonical] = chance.get(canonical, 0.0) + amt * count

    return guaranteed, chance


# ── Change Detection ──────────────────────────────────────────────────────────

def get_previous_production(history: dict, city_name: str, run_date: str) -> dict:
    """
    Return the most recent production snapshot for a city BEFORE today.

    Returns:
        {kit_name: frags_per_day}  — empty dict if no prior history
    """
    city_history = history.get(city_name, {})
    prev = {}
    for kit, entries in city_history.items():
        # entries: [[date, value], ...]
        prior = [e for e in entries if e[0] < run_date]
        if prior:
            # Most recent prior entry
            latest = max(prior, key=lambda e: e[0])
            prev[kit] = latest[1]
    return prev


def compute_change(current: float, previous: dict, kit_name: str) -> str:
    """
    Compare current production to previous run.

    Returns a human-readable change string:
        'NEW'       — kit not in previous run
        'GONE'      — (not used here; handled in caller)
        '+X.X'      — production increased by >= threshold
        '-X.X'      — production decreased by >= threshold
        ''          — no meaningful change
    """
    if kit_name not in previous:
        return "NEW"
    delta = current - previous[kit_name]
    if abs(delta) < CHANGE_THRESHOLD:
        return ""
    sign = "+" if delta > 0 else ""
    return f"{sign}{delta:.1f}"


def get_gone_kits(current_production: dict, previous: dict) -> list[str]:
    """Return kit names that were in the previous run but not in current."""
    return [k for k in previous if k not in current_production]


# ── Dashboard printing ────────────────────────────────────────────────────────

def build_dashboard(
    production  : dict[str, float],
    chance_prod : dict[str, float],
    frags_needed: dict[str, tuple[str, int]],
    banked      : dict[str, float],
    city_name   : str,
    previous    : dict = None,
):
    """Print the fragment dashboard with optional change column."""

    if previous is None:
        previous = {}

    run_date = date.today().strftime("%Y-%m-%d")

    rows = []
    for canonical, daily in production.items():
        upper = clean_name(canonical)
        if upper in frags_needed:
            display_name, needed = frags_needed[upper]
        else:
            display_name = canonical
            needed       = None

        bank = banked.get(canonical, 0.0)

        if needed is not None and daily > 0:
            remaining = max(0.0, needed - bank)
            days      = remaining / daily
        else:
            remaining = None
            days      = None

        change = compute_change(daily, previous, canonical)

        rows.append({
            "Kit"       : display_name,
            "Canonical" : canonical,
            "Frags/Day" : daily,
            "Need"      : needed if needed is not None else "?",
            "Banked"    : bank,
            "Remaining" : remaining if remaining is not None else "?",
            "Days Left" : days,
            "Change"    : change,
        })

    rows.sort(key=lambda r: (r["Days Left"] is None, r["Days Left"] or 0))

    col_kit = max(len(r["Kit"]) for r in rows) + 2
    col_kit = max(col_kit, 35)

    header = (
        f"{'Kit':<{col_kit}}"
        f"{'Frags/Day':>10}"
        f"{'△ Change':>10}"
        f"{'Need':>6}"
        f"{'Banked':>8}"
        f"{'Remaining':>10}"
        f"{'Days Left':>10}"
    )
    sep = "─" * len(header)

    print()
    print("=" * len(header))
    print(f"  FoE FRAGMENT DASHBOARD  ·  City: {city_name}")
    print("=" * len(header))
    print(header)
    print(sep)

    for r in rows:
        days_str   = f"{r['Days Left']:.1f}" if r["Days Left"] is not None else "  N/A"
        rem_str    = f"{r['Remaining']:.0f}" if isinstance(r["Remaining"], float) else str(r["Remaining"])
        change_str = r["Change"] if r["Change"] else "—"
        print(
            f"{r['Kit']:<{col_kit}}"
            f"{r['Frags/Day']:>10.1f}"
            f"{change_str:>10}"
            f"{str(r['Need']):>6}"
            f"{r['Banked']:>8.0f}"
            f"{rem_str:>10}"
            f"{days_str:>10}"
        )

    print(sep)
    print()

    # Gone kits (were producing, now aren't)
    gone = get_gone_kits(production, previous)
    if gone:
        print("  ⬇  Kits no longer being produced (vs last run):")
        for k in gone:
            old_val = previous[k]
            print(f"     • {k}  (was {old_val:.1f}/day)")
        print()

    unknown = [r["Kit"] for r in rows if r["Need"] == "?"]
    if unknown:
        print("  ⚠  Kits produced but not found in LUT_Item_FragmentsNeeded:")
        for k in unknown:
            print(f"     • {k}")
        print()

    produced_upper = {k.upper() for k in production}
    not_produced = [
        display for upper, (display, _) in frags_needed.items()
        if upper not in produced_upper
    ]
    if not_produced:
        print("  ℹ  Kits in your LUT but not currently produced by any building:")
        for k in sorted(not_produced):
            print(f"     • {k}")
        print()

    # Chance-based section
    if chance_prod:
        chance_rows = []
        for canonical, avg_daily in chance_prod.items():
            upper = clean_name(canonical)
            if upper in frags_needed:
                display_name, needed = frags_needed[upper]
                avg_days = needed / avg_daily if avg_daily > 0 else None
            else:
                display_name = canonical
                needed       = None
                avg_days     = None

            change = compute_change(avg_daily, previous, canonical)

            chance_rows.append({
                "Kit"      : display_name,
                "Canonical": canonical,
                "Avg/Day"  : avg_daily,
                "Need"     : needed,
                "Avg Days" : avg_days,
                "Change"   : change,
            })

        chance_rows.sort(key=lambda r: (r["Avg Days"] is None, r["Avg Days"] or 0))

        col_kit = max(len(r["Kit"]) for r in chance_rows) + 2
        col_kit = max(col_kit, 35)

        header2 = (
            f"{'Kit':<{col_kit}}"
            f"{'Avg Frags/Day':>14}"
            f"{'△ Change':>10}"
            f"{'Need':>6}"
            f"{'Avg Days/Kit':>14}"
        )
        sep2 = "─" * len(header2)

        print()
        print("=" * len(header2))
        print(f"  CHANCE-BASED PRODUCTIONS (daily averages)  ·  City: {city_name}")
        print("=" * len(header2))
        print(header2)
        print(sep2)

        for r in chance_rows:
            days_str   = f"{r['Avg Days']:.1f}" if r["Avg Days"] is not None else "  N/A"
            need_str   = str(r["Need"]) if r["Need"] is not None else "?"
            change_str = r["Change"] if r["Change"] else "—"
            print(
                f"{r['Kit']:<{col_kit}}"
                f"{r['Avg/Day']:>14.1f}"
                f"{change_str:>10}"
                f"{need_str:>6}"
                f"{days_str:>14}"
            )
        print(sep2)
        print()


# ── Shared row builder ────────────────────────────────────────────────────────

def build_rows(production, frags_needed, banked, previous=None):
    """Build sorted list of dashboard rows shared by print and Excel/HTML outputs."""
    if previous is None:
        previous = {}

    rows = []
    for canonical, daily in production.items():
        upper = clean_name(canonical)
        if upper in frags_needed:
            display_name, needed = frags_needed[upper]
        else:
            display_name = canonical
            needed       = None

        bank = banked.get(canonical, 0.0)

        if needed is not None and daily > 0:
            remaining = max(0.0, needed - bank)
            days      = remaining / daily
        else:
            remaining = None
            days      = None

        change = compute_change(daily, previous, canonical)

        rows.append({
            "Kit"       : display_name,
            "Canonical" : canonical,
            "Frags/Day" : daily,
            "Need"      : needed,
            "Days Left" : days,
            "Change"    : change,
        })

    rows.sort(key=lambda r: (r["Days Left"] is None, r["Days Left"] or 0))
    return rows


# ── History tracking ──────────────────────────────────────────────────────────

def load_history() -> dict:
    """Load existing history or return empty structure."""
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_history(history: dict):
    """Save history to JSON file."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=2, ensure_ascii=False)


def update_history(history: dict, city_name: str, production: dict, run_date: str):
    """Add today's production snapshot to the history."""
    if city_name not in history:
        history[city_name] = {}

    for kit, daily in production.items():
        if kit not in history[city_name]:
            history[city_name][kit] = []

        dates = [e[0] for e in history[city_name][kit]]
        if run_date not in dates:
            history[city_name][kit].append([run_date, daily])
        else:
            for entry in history[city_name][kit]:
                if entry[0] == run_date:
                    entry[1] = daily
                    break

    return history


def get_all_kits(history: dict) -> list[str]:
    """Return sorted list of all kit names across all cities."""
    kits = set()
    for city_data in history.values():
        kits.update(city_data.keys())
    return sorted(kits)


def get_all_cities(history: dict) -> list[str]:
    """Return sorted list of all city names in history."""
    return sorted(history.keys())


# ── HTML export ───────────────────────────────────────────────────────────────

def _change_badge(change: str) -> str:
    """Return an HTML badge for the change value."""
    if not change:
        return '<span class="chg chg-none">—</span>'
    if change == "NEW":
        return '<span class="chg chg-new">NEW</span>'
    if change == "GONE":
        return '<span class="chg chg-gone">GONE</span>'
    if change.startswith("+"):
        return f'<span class="chg chg-up">{change}</span>'
    return f'<span class="chg chg-down">{change}</span>'


def generate_html(city_data: list[tuple], run_date: str) -> str:
    """
    Generate a self-contained HTML dashboard page.
    city_data: list of (city_display, guaranteed_rows, chance_rows, gone_kits)
    """

    def table_rows_guaranteed(rows):
        html = ""
        for r in rows:
            days = f"{r['Days Left']:.1f}" if r['Days Left'] is not None else "N/A"
            need = str(r['Need']) if r['Need'] is not None else "?"
            kpm  = f"{30 / r['Days Left']:.2f}" if r['Days Left'] else "N/A"
            cls  = "unknown" if r['Need'] is None else ""
            badge = _change_badge(r.get("Change", ""))
            html += f"""
            <tr class="{cls}">
                <td>{r['Kit']}</td>
                <td class="num">{r['Frags/Day']:.1f}</td>
                <td class="num chg-cell">{badge}</td>
                <td class="num">{need}</td>
                <td class="num">{days}</td>
                <td class="num">{kpm}</td>
            </tr>"""
        return html

    def table_rows_chance(rows):
        html = ""
        for r in rows:
            days  = f"{r['Avg Days']:.1f}" if r['Avg Days'] is not None else "N/A"
            need  = str(r['Need']) if r['Need'] is not None else "?"
            cls   = "unknown" if r['Need'] is None else ""
            badge = _change_badge(r.get("Change", ""))
            html += f"""
            <tr class="{cls}">
                <td>{r['Kit']}</td>
                <td class="num">{r['Avg/Day']:.1f}</td>
                <td class="num chg-cell">{badge}</td>
                <td class="num">{need}</td>
                <td class="num">{days}</td>
            </tr>"""
        return html

    def gone_section(gone_kits):
        if not gone_kits:
            return ""
        items = "".join(f'<li>{k} <span class="was">(was {v:.1f}/day)</span></li>' for k, v in gone_kits)
        return f"""
        <div class="gone-block">
            <span class="gone-label">⬇ No Longer Producing</span>
            <ul>{items}</ul>
        </div>"""

    city_sections = ""
    for city_display, guar_rows, chance_rows, gone_kits in city_data:
        city_sections += f"""
        <section class="city">
            <h2>{city_display}</h2>
            {gone_section(gone_kits)}

            <h3>⚔ Guaranteed Daily Production</h3>
            <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Kit</th>
                        <th>Frags / Day</th>
                        <th>△ Change</th>
                        <th>Need</th>
                        <th>Days / Kit</th>
                        <th>Kits / Month</th>
                    </tr>
                </thead>
                <tbody>{table_rows_guaranteed(guar_rows)}</tbody>
            </table>
            </div>

            <h3>🎲 Chance-Based (Daily Averages)</h3>
            <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Kit</th>
                        <th>Avg Frags / Day</th>
                        <th>△ Change</th>
                        <th>Need</th>
                        <th>Avg Days / Kit</th>
                    </tr>
                </thead>
                <tbody>{table_rows_chance(chance_rows)}</tbody>
            </table>
            </div>
        </section>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>FoE Fragment Dashboard</title>
<link href="https://fonts.googleapis.com/css2?family=Cinzel:wght@400;700&family=Crimson+Text:ital,wght@0,400;0,600;1,400&display=swap" rel="stylesheet">
<style>
  :root {{
    --bg-deep:    #0d0a07;
    --bg-card:    #1a1208;
    --gold:       #c9a84c;
    --gold-lt:    #e8c96a;
    --gold-dim:   #7a6530;
    --parchment:  #e8d9b0;
    --parch-dim:  #a89060;
    --border:     #3d2e10;
    --border-gold:#5a4520;
    --shadow:     rgba(0,0,0,0.7);
    --chg-up:     #8fcc8f;
    --chg-down:   #cc7070;
    --chg-new-bg: rgba(42,92,42,0.3);
    --chg-new-fg: #8fcc8f;
    --chg-gone-bg:rgba(139,32,32,0.3);
    --chg-gone-fg:#cc7070;
  }}

  * {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    background-color: var(--bg-deep);
    background-image:
      radial-gradient(ellipse at 20% 10%, rgba(201,168,76,0.07) 0%, transparent 50%),
      radial-gradient(ellipse at 80% 90%, rgba(139,80,20,0.07) 0%, transparent 50%);
    color: var(--parchment);
    font-family: 'Crimson Text', Georgia, serif;
    font-size: 16px;
    line-height: 1.6;
    min-height: 100vh;
  }}

  header {{
    background: linear-gradient(135deg, #2a1e08 0%, #1a1208 100%);
    border-bottom: 2px solid var(--gold-dim);
    padding: 2rem 2rem 1.5rem;
    text-align: center;
    position: relative;
  }}
  header::before {{
    content: "⚜"; position: absolute; left: 2rem; top: 50%;
    transform: translateY(-50%); font-size: 2rem; color: var(--gold); opacity: 0.5;
  }}
  header::after {{
    content: "⚜"; position: absolute; right: 2rem; top: 50%;
    transform: translateY(-50%); font-size: 2rem; color: var(--gold); opacity: 0.5;
  }}

  h1 {{
    font-family: 'Cinzel', serif;
    font-size: clamp(1.4rem, 4vw, 2.2rem);
    color: var(--gold-lt);
    letter-spacing: 0.08em;
    text-shadow: 0 0 30px rgba(201,168,76,0.3);
  }}
  .subtitle {{
    font-size: 0.95rem; color: var(--parch-dim);
    margin-top: 0.3rem; font-style: italic;
  }}

  main {{ max-width: 1100px; margin: 0 auto; padding: 2rem 1rem; }}

  .city {{
    margin-bottom: 3rem;
    background: var(--bg-card);
    border: 1px solid var(--border-gold);
    border-radius: 8px;
    overflow: hidden;
    box-shadow: 0 8px 32px var(--shadow), inset 0 1px 0 rgba(201,168,76,0.08);
  }}

  h2 {{
    font-family: 'Cinzel', serif; font-size: 1.4rem; color: var(--gold);
    background: linear-gradient(135deg, #2a1e08, #1a1208);
    padding: 1rem 1.5rem;
    border-bottom: 1px solid var(--border-gold);
    letter-spacing: 0.06em;
  }}

  h3 {{
    font-family: 'Cinzel', serif; font-size: 0.85rem; color: var(--gold-dim);
    padding: 1rem 1.5rem 0.5rem; letter-spacing: 0.12em;
    text-transform: uppercase;
  }}

  /* Gone block */
  .gone-block {{
    margin: 0.75rem 1.5rem;
    padding: 0.6rem 1rem;
    background: var(--chg-gone-bg);
    border: 1px solid #6a3a3a;
    border-radius: 5px;
    font-size: 0.9rem;
  }}
  .gone-label {{
    font-family: 'Cinzel', serif; font-size: 0.8rem;
    color: var(--chg-gone-fg); letter-spacing: 0.08em;
    text-transform: uppercase; display: block; margin-bottom: 0.3rem;
  }}
  .gone-block ul {{ list-style: none; padding: 0; }}
  .gone-block li {{ color: var(--parchment); padding: 0.1rem 0; }}
  .was {{ color: var(--parch-dim); font-style: italic; font-size: 0.85em; }}

  .table-wrap {{ overflow-x: auto; padding: 0 1rem 1.2rem; }}

  table {{ width: 100%; border-collapse: collapse; font-size: 0.95rem; }}

  thead tr {{
    background: linear-gradient(135deg, #2a1e08, #3d2e10);
    border-bottom: 2px solid var(--border-gold);
  }}
  th {{
    font-family: 'Cinzel', serif; font-size: 0.75rem; font-weight: 700;
    color: var(--gold-lt); letter-spacing: 0.08em; text-transform: uppercase;
    padding: 0.7rem 1rem; text-align: left; white-space: nowrap;
  }}
  th.num, td.num {{ text-align: right; }}

  tbody tr {{
    border-bottom: 1px solid var(--border);
    transition: background 0.15s;
  }}
  tbody tr:hover {{ background: rgba(201,168,76,0.06); }}
  tbody tr:nth-child(even) {{ background: rgba(0,0,0,0.2); }}
  tbody tr:nth-child(even):hover {{ background: rgba(201,168,76,0.06); }}
  tbody tr.unknown {{ opacity: 0.55; font-style: italic; }}

  td {{ padding: 0.55rem 1rem; color: var(--parchment); }}
  td:first-child {{ color: var(--parchment); }}
  td.num {{ color: var(--gold-lt); font-variant-numeric: tabular-nums; }}

  /* Change badges */
  .chg-cell {{ text-align: right; }}
  .chg {{
    display: inline-block;
    font-family: 'Cinzel', serif;
    font-size: 0.75rem;
    font-weight: 700;
    letter-spacing: 0.04em;
    padding: 0.1rem 0.45rem;
    border-radius: 3px;
    min-width: 3.5rem;
    text-align: center;
  }}
  .chg-none  {{ color: var(--border-gold); background: transparent; font-weight: 400; }}
  .chg-up    {{ color: var(--chg-up);      background: rgba(42,92,42,0.3);  border: 1px solid #3a6a3a; }}
  .chg-down  {{ color: var(--chg-down);    background: rgba(139,32,32,0.3); border: 1px solid #6a3a3a; }}
  .chg-new   {{ color: var(--chg-new-fg);  background: var(--chg-new-bg);   border: 1px solid #3a6a3a; }}
  .chg-gone  {{ color: var(--chg-gone-fg); background: var(--chg-gone-bg);  border: 1px solid #6a3a3a; }}

  footer {{
    text-align: center; padding: 1.5rem; color: var(--parch-dim);
    opacity: 0.5; font-size: 0.85rem; font-style: italic;
    border-top: 1px solid var(--border);
  }}

  @media (max-width: 600px) {{
    header::before, header::after {{ display: none; }}
    th, td {{ padding: 0.5rem 0.6rem; font-size: 0.85rem; }}
  }}
</style>
</head>
<body>
<header>
  <h1>⚔ FoE Fragment Dashboard ⚔</h1>
  <div class="subtitle">Updated {run_date}</div>
</header>
<main>
{city_sections}
</main>
<footer>Forge of Empires Fragment Dashboard &nbsp;·&nbsp; Generated by foe_dashboard.py</footer>
</body>
</html>"""


def write_html(city_data: list[tuple], output_path: str):
    """Write the HTML dashboard to a file."""
    run_date = date.today().strftime("%d %b %Y")
    html = generate_html(city_data, run_date)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  HTML dashboard saved: {output_path}")


# ── Excel export ──────────────────────────────────────────────────────────────

def write_excel(city_data, output_path, history=None):
    """Write dashboard for one or more cities to an Excel workbook."""
    wb = Workbook()
    wb.remove(wb.active)

    COL_HEADER_BG = "1F4E79"
    COL_HEADER_FG = "FFFFFF"
    COL_TITLE_BG  = "2E75B6"
    COL_ALT_ROW   = "DEEAF1"
    COL_UNKNOWN   = "FFF2CC"
    COL_BORDER    = "BDD7EE"

    thin   = Side(style="thin", color=COL_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for city_name, rows in city_data:
        ws = wb.create_sheet(title=city_name[:31])

        ws.merge_cells("A1:E1")
        t = ws["A1"]
        t.value     = f"FoE Fragment Dashboard  ·  {city_name}  ·  {date.today().strftime('%d %b %Y')}"
        t.font      = Font(name="Arial", bold=True, size=13, color=COL_HEADER_FG)
        t.fill      = PatternFill("solid", fgColor=COL_TITLE_BG)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for col, hdr in enumerate(["Kit", "Frags / Day", "Frags Needed", "Days Left", "Kits / Month"], 1):
            c = ws.cell(row=2, column=col, value=hdr)
            c.font      = Font(name="Arial", bold=True, color=COL_HEADER_FG)
            c.fill      = PatternFill("solid", fgColor=COL_HEADER_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border
        ws.row_dimensions[2].height = 18

        for i, r in enumerate(rows):
            rn  = i + 3
            bg  = COL_UNKNOWN if r["Need"] is None else (COL_ALT_ROW if i % 2 else "FFFFFF")
            fil = PatternFill("solid", fgColor=bg)

            def cell(col, val, fmt=None, center=False):
                c = ws.cell(row=rn, column=col, value=val)
                c.font      = Font(name="Arial", size=10)
                c.fill      = fil
                c.border    = border
                if center:
                    c.alignment = Alignment(horizontal="center")
                if fmt:
                    c.number_format = fmt
                return c

            cell(1, r["Kit"])
            cell(2, r["Frags/Day"], fmt="0.0",  center=True)
            cell(3, r["Need"] if r["Need"] is not None else "?", center=True)
            cell(4, f"=C{rn}/B{rn}" if r["Need"] is not None else "N/A", fmt="0.0", center=True)
            cell(5, f"=B{rn}*30/C{rn}" if r["Need"] is not None else "N/A", fmt="0.00", center=True)

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 14
        ws.freeze_panes = "A3"

    if history:
        _write_history_sheet(wb, history)

    wb.save(output_path)
    print(f"  Excel workbook saved: {output_path}")


def _write_history_sheet(wb, history: dict):
    """Add a History sheet with raw data table for charting."""
    from openpyxl.chart import LineChart, Reference
    from openpyxl.worksheet.datavalidation import DataValidation

    COL_HEADER_BG = "1F4E79"
    COL_HEADER_FG = "FFFFFF"
    COL_TITLE_BG  = "2E75B6"
    COL_BORDER    = "BDD7EE"
    thin   = Side(style="thin", color=COL_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cities   = get_all_cities(history)
    all_kits = get_all_kits(history)

    all_dates = sorted({
        entry[0]
        for city_data in history.values()
        for entries in city_data.values()
        for entry in entries
    })

    if not all_dates:
        return

    raw = wb.create_sheet(title="_RawHistory")
    raw.sheet_state = "hidden"

    raw.cell(row=1, column=1, value="Date")
    col = 2
    col_map = {}
    for city in cities:
        for kit in all_kits:
            if kit in history.get(city, {}):
                raw.cell(row=1, column=col, value=f"{city} | {kit}")
                col_map[(city, kit)] = col
                col += 1
    total_cols = col - 1

    for r, dt in enumerate(all_dates, 2):
        raw.cell(row=r, column=1, value=dt)
        for (city, kit), c in col_map.items():
            val = None
            for entry in history.get(city, {}).get(kit, []):
                if entry[0] == dt:
                    val = entry[1]
                    break
            if val is not None:
                raw.cell(row=r, column=c, value=val)
    total_rows = len(all_dates) + 1

    ws = wb.create_sheet(title="History", index=0)

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value     = f"FoE Fragment Production History  ·  {date.today().strftime('%d %b %Y')}"
    t.font      = Font(name="Arial", bold=True, size=13, color=COL_HEADER_FG)
    t.fill      = PatternFill("solid", fgColor=COL_TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws["A3"] = "Select Kit:"
    ws["A3"].font = Font(name="Arial", bold=True, size=11)
    ws["B3"] = all_kits[0] if all_kits else ""
    ws["B3"].font = Font(name="Arial", size=11)
    ws["B3"].fill = PatternFill("solid", fgColor="DEEAF1")

    for i, kit in enumerate(all_kits, 1):
        raw.cell(row=total_rows + 2 + i, column=1, value=kit)
    kit_start = total_rows + 3
    kit_end   = total_rows + 2 + len(all_kits)

    dv2 = DataValidation(
        type="list",
        formula1=f"_RawHistory!$A${kit_start}:$A${kit_end}",
        showDropDown=False,
        allow_blank=False
    )
    ws.add_data_validation(dv2)
    dv2.add(ws["B3"])

    ws["A5"] = "Date"
    ws["A5"].font  = Font(name="Arial", bold=True, color=COL_HEADER_FG)
    ws["A5"].fill  = PatternFill("solid", fgColor=COL_HEADER_BG)
    ws["A5"].border = border

    for ci, city in enumerate(cities, 1):
        c = ws.cell(row=5, column=ci + 1, value=city)
        c.font  = Font(name="Arial", bold=True, color=COL_HEADER_FG)
        c.fill  = PatternFill("solid", fgColor=COL_HEADER_BG)
        c.border = border

    for ri, dt in enumerate(all_dates, 6):
        ws.cell(row=ri, column=1, value=dt).border = border
        for ci, city in enumerate(cities, 1):
            formula = (
                f'=IFERROR(INDEX(_RawHistory!$A$2:${get_column_letter(total_cols)}${total_rows},'
                f'MATCH($A{ri},_RawHistory!$A$2:$A${total_rows},0),'
                f'MATCH("{city} | "&$B$3,_RawHistory!$A$1:${get_column_letter(total_cols)}$1,0)),"")'
            )
            c = ws.cell(row=ri, column=ci + 1, value=formula)
            c.number_format = "0.0"
            c.border = border

    ws.column_dimensions["A"].width = 14
    for i in range(len(cities)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16

    last_data_row = 5 + len(all_dates)
    chart = LineChart()
    chart.title        = "Fragment Production History"
    chart.style        = 10
    chart.y_axis.title = "Frags / Day"
    chart.x_axis.title = "Date"
    chart.width        = 22
    chart.height       = 14

    for ci, city in enumerate(cities, 1):
        data_ref = Reference(ws, min_col=ci + 1, max_col=ci + 1,
                             min_row=5, max_row=last_data_row)
        chart.add_data(data_ref, titles_from_data=True)

    dates_ref = Reference(ws, min_col=1, min_row=6, max_row=last_data_row)
    chart.set_categories(dates_ref)

    for series in chart.series:
        series.smooth = True

    ws.add_chart(chart, f"A{last_data_row + 3}")
    ws.freeze_panes = "A6"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        sys.exit(1)

    export_excel = "--excel" in args
    json_files   = [a for a in args if not a.startswith("--")]

    if not json_files:
        print("Error: no city JSON files specified.")
        sys.exit(1)

    for f in json_files:
        if not os.path.exists(f):
            print(f"Error: file not found: {f}")
            sys.exit(1)

    print("Loading lookup tables...")
    frags_needed, kit_mapping = load_luts()
    city_names = load_city_names()
    print(f"  {len(frags_needed)} kits in fragment LUT.")
    print(f"  {len(kit_mapping)} entries in kit mapping LUT.")
    print(f"  {len(city_names)} city names loaded.")

    banked     = {}
    excel_data = []
    html_data  = []
    history    = load_history()
    run_date   = date.today().strftime("%Y-%m-%d")

    for json_path in json_files:
        stem   = Path(json_path).stem
        parts  = stem.split("_")
        city_abbrev = stem
        for part in parts:
            if part not in ("LUT", "Buildings") and not part.isdigit() and len(part) <= 8:
                city_abbrev = part
                break

        city_name    = city_abbrev
        city_display = city_names.get(city_abbrev, city_abbrev)

        print(f"\nLoading city: {json_path}")
        city = load_city(json_path)
        print(f"  {len(city)} buildings loaded.")

        # Load previous production BEFORE updating history
        previous = get_previous_production(history, city_name, run_date)
        if previous:
            print(f"  Comparing against previous run ({max(e[0] for kit in history.get(city_name, {}).values() for e in kit if e[0] < run_date)}).")
        else:
            print("  No prior history found — change detection will show NEW for all kits.")

        print("Calculating daily fragment production...")
        production, chance_prod = collect_fragment_production(city, kit_mapping)
        print(f"  {len(production)} kits being produced (guaranteed).")
        print(f"  {len(chance_prod)} kits being produced (chance-based).")

        rows = build_rows(production, frags_needed, banked, previous=previous)
        excel_data.append((city_display, rows))

        # Build chance rows for HTML
        chance_rows = []
        for canonical, avg_daily in chance_prod.items():
            upper = clean_name(canonical)
            if upper in frags_needed:
                display_name, needed = frags_needed[upper]
                avg_days = needed / avg_daily if avg_daily > 0 else None
            else:
                display_name = canonical
                needed       = None
                avg_days     = None
            change = compute_change(avg_daily, previous, canonical)
            chance_rows.append({
                "Kit": display_name, "Avg/Day": avg_daily,
                "Need": needed, "Avg Days": avg_days, "Change": change,
            })
        chance_rows.sort(key=lambda r: (r["Avg Days"] is None, r["Avg Days"] or 0))

        # Gone kits for HTML
        gone_kits = [(k, previous[k]) for k in get_gone_kits(production, previous)]

        html_data.append((city_display, rows, chance_rows, gone_kits))

        # Update history AFTER comparison
        update_history(history, city_name, production, run_date)

        # Console output
        build_dashboard(production, chance_prod, frags_needed, banked, city_display, previous=previous)

    save_history(history)
    print(f"  History updated: {HISTORY_FILE}")

    OUTPUT_DIR.mkdir(exist_ok=True)
    html_path = OUTPUT_DIR / "index.html"
    write_html(html_data, str(html_path))

    if export_excel:
        today    = date.today().strftime("%Y%m%d")
        out_path = OUTPUT_DIR / f"FoE_Dashboard_{today}.xlsx"
        write_excel(excel_data, str(out_path), history=history)


if __name__ == "__main__":
    main()
